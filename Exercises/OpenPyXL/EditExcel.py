import openpyxl
wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'].value == None
sheet['A1'].value = '42'
sheet['A2'].value = 'Hello World'
import os
os.chdir('c:\\users\\alan\\documents')
wb.save('example2.xlsx')
"""openpyxl.load_workbook('fasfasgasf.xlsx')"""
sheet2= wb.create_sheet()
wb.get_sheet_names()
sheet2.title = 'My New Sheet'
wb.get_sheet_names()
wb.save('example3.xlsx')
sheet3= wb.create_sheet(index=0, title='My Primary Sheet')
wb.save('example4.xlsx')