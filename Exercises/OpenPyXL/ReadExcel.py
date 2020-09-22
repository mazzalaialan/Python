import openpyxl
import os
os.chdir('c:\\users\\alan\\documents')
workbook = openpyxl.load_workbook('example.xlsx')
type(workbook)
workbook.get_sheet_names()
sheet = workbook.get_sheet_by_name('Sheet1')
cell = sheet['A1']
cell.value
print(str(cell.value))
print(str(sheet['A1'].value))
print(sheet.cell(row=1, column=1).value)
for i in range(1,8):
    print(i,sheet.cell(row=i, column=2).value)